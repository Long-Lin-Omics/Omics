import pandas as pd
import numpy as np
import sys
import io
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import re


matrix_data = sys.argv[1]
output_excel = sys.argv[2]
n_samples = int(sys.argv[3])
apply_log = len(sys.argv) > 4 and sys.argv[4] == 'log'
# a=int(sys.argv[4])
# b=int(sys.argv[5])
# bin_size=int(sys.argv[6])

# matrix_data = 'test.matrix.data.tsv'
# Read the first three lines
with open( matrix_data, 'r') as file:
    lines = file.readlines()
    header_line = lines[2].strip().split('\t')
    n_genes = int(re.search(r'genes:(\d+)', lines[0]).group(1))
    b = int(re.search(r'downstream:(\d+)', lines[1]).group(1))
    a = int(re.search(r'upstream:(\d+)', lines[1]).group(1))*-1
    bin_size = int(re.search(r'bin size:(\d+)', lines[1]).group(1))


# Load matrix data
df = pd.read_csv(matrix_data, sep='\t', skiprows=3, header=None) 

df_new = pd.DataFrame(df.sum(axis=0).values.reshape(n_samples,-1)).T / n_genes
df_new.columns = [header_line[int(df.shape[1]/n_samples)*i+1] for i in range(n_samples)] 

if apply_log:
    print("Applying log2 transformation...")
    df_new = np.log2(df_new + 0.01)

Bin_start = [i for i in range(a,b,bin_size)]
Bin_end = [i + bin_size -1 for i in range(a,b,bin_size)]

df_new['Bin_start'] = Bin_start
df_new['Bin_end'] = Bin_end

# Reorder the columns if needed (put 'Bin_start' and 'Bin_end' at the front)
df_new = df_new[['Bin_start', 'Bin_end'] + [col for col in df_new.columns if col not in ['Bin_start', 'Bin_end']]]
print(df_new)

# Create a new Excel workbook
wb = Workbook()

# Create a new sheet called 'data' and write the DataFrame to it
ws_data = wb.create_sheet("data")
for r in dataframe_to_rows(df_new, index=False, header=True):
    ws_data.append(r)

                # # Create the metaplot and save it as a PNG file
                # plt.figure(figsize=(8, 5))
                # plt.plot(df_new.iloc[:, 0], df_new.iloc[:, 2:], linewidth=2)
                # plt.xlabel("Genomic Position")
                # ylabel = "Log2 Signal Intensity" if apply_log else "Signal Intensity"
                # plt.ylabel(ylabel)
                # title = "Metaplot (log2-transformed)" if apply_log else "Metaplot"
                # plt.title(title)
                # plt.legend(df_new.columns[2:], title="Samples")
                # plt.grid()

                # # Save plot to a temporary file
                # plot_file = output_excel + ".metaplot.png"
                # plt.savefig(plot_file)
                # plt.close()  # Free memory

plt.figure(figsize=(8, 5))

x = df_new.iloc[:, 0].values
sample_cols = list(df_new.columns[2:])  # samples start from col index 2

# A set of distinct marker shapes (循环使用，样本多也不会报错)
markers = ['o', 's', '^', 'D', 'v', 'P', 'X', '*', '<', '>', 'h', 'H', '+', 'x', '1', '2', '3', '4']

# 让 marker 不至于太密：每隔若干点打一个
# 你也可以改成更大/更小，比如 10 或 30
markevery = max(1, len(x) // 25)

for i, col in enumerate(sample_cols):
    y = df_new[col].values
    m = markers[i % len(markers)]
    plt.plot(
        x, y,
        linewidth=2,
        marker=m,
        markevery=markevery,
        markersize=6,
        markerfacecolor='none',   # 空心点更清晰（也更适合黑白打印）
        markeredgewidth=1.5,
        label=col
    )

plt.xlabel("Genomic Position")
ylabel = "Log2 Signal Intensity" if apply_log else "Signal Intensity"
plt.ylabel(ylabel)
title = "Metaplot (log2-transformed)" if apply_log else "Metaplot"
plt.title(title)
plt.legend(title="Samples", ncol=1, fontsize=9)  # 样本很多可调 ncol=2/3
plt.grid(True)

# Save plot to a temporary file
plot_file = output_excel + ".metaplot.png"
plt.savefig(plot_file, dpi=200, bbox_inches="tight")
plt.close()

# Add the metaplot image to a new sheet called 'Metaplot'
ws_metaplot = wb.create_sheet("Metaplot")
img = Image(plot_file)
ws_metaplot.add_image(img, "B2")

# Remove the default sheet created
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# Save the new Excel file with both the data and the image
wb.save(output_excel)

print("New Excel file created with data and metaplot image!")