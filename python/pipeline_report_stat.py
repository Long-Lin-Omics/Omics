import pandas as pd
import argparse
import os
import re
import sys
import gzip
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def extract_read_count(fastq_file):
    """Extracts read count from a FASTQ file (plain or gzipped) by counting the lines and dividing by 4."""
    open_func = gzip.open if fastq_file.endswith(".gz") else open
    print("dealing with fastq: " + fastq_file)
    with open_func(fastq_file, 'rt') as f:  # 'rt' mode ensures reading as text
        return sum(1 for _ in f) // 4

def extract_read_count_report(fastq_report_file):
    with open(fastq_report_file, 'rt') as f:  # 'rt' mode ensures reading as text
        return int(f.readline().strip())


def parse_duplication_metrics(metrics_file):
    """Parses Picard MarkDuplicates metrics.txt file and extracts duplication metrics."""
    with open(metrics_file, 'r') as f:
        lines = f.readlines()
    
    start_idx = next(i for i, line in enumerate(lines) if line.startswith('LIBRARY')) + 1
    values = lines[start_idx].split('\t')
    
    return {
        "Unpaired Reads Examined": int(values[1]),
        "Read Pairs Examined": int(values[2]),
        "Secondary or supplementary RDS": int(values[3]) ,  
        "Unmapped Reads": int(values[4]),
        "Unpaired Read Duplicates": int(values[5]),        
        "Read Pair Duplicates": int(values[6]),
        "Percent Duplication": float(values[8])
          }

def count_peaks(peak_file):
    """Counts the number of peaks in a MACS2 peak file."""
    with open(peak_file, 'r') as f:
        return sum(1 for _ in f)

import numpy as np

def calculate_peak_length_quantiles(peak_file, probs=[0, 0.25, 0.5, 0.75,1]):
    """
    Calculates quantiles of peak lengths from a MACS2 peak file.
    Parameters:
        peak_file (str): Path to the peak file (.narrowPeak, .broadPeak, etc.)
        probs (list of float): List of quantiles to calculate (e.g., [0.25, 0.5, 0.75])
    Returns:
        dict: Dictionary of {quantile: value}
    """
    peak_lengths = []
    with open(peak_file) as f:
        for line in f:
            fields = line.strip().split('\t')
            if len(fields) >= 3:
                try:
                    start = int(fields[1])
                    end = int(fields[2])
                    peak_lengths.append(end - start)
                except ValueError:
                    continue  # skip malformed lines
    if not peak_lengths:
        raise ValueError("No valid peaks found in the file.")
    quantiles = np.quantile(peak_lengths, probs)
    return dict(zip(probs, quantiles))

def find_matching_file_or_by_order(sample, file_list):
    """Finds a matching file from the list that contains the sample name or uses the same order."""
    matches = [f for f in file_list if sample in f]
    to_return = min(matches, key=len) if matches else file_list.pop(0) if file_list else None
    if (not to_return == None) and (not os.path.exists(to_return)):
        print(f"Error: File '{to_return}' does not exist.") 
        sys.exit(1) 
    return to_return

def find_matching_file(sample, file_list):
    """Finds a matching file from the list that contains the sample name or uses the same order."""
    matches = [f for f in file_list if sample in f]
    to_return = matches[0] if matches else None
    if (not to_return == None) and (not os.path.exists(to_return)):
        print(f"Error: File '{to_return}' does not exist.") 
        sys.exit(1) 
    return to_return


def main():
    parser = argparse.ArgumentParser(description="Summarize sequencing metrics from multiple sources.")
    parser.add_argument("--fastq", nargs='+', help="List of raw FASTQ files",default=[])
    parser.add_argument("--trimmed_fastq", nargs='+', help="List of trimmed FASTQ files", default=[])
    parser.add_argument("--metrics", nargs='+', help="List of Picard MarkDuplicates metrics.txt files", default=[])
    parser.add_argument("--peaks", nargs='+', help="List of MACS2 peak files", default=[])
    parser.add_argument("--samples", nargs='+', required=True, help="List of sample names")
    parser.add_argument("--output", default="summary.tsv", help="Output TSV file")
    parser.add_argument("--fastq_report",action='store_true', help='Input fastq report list instead of fastq from --fastq')
    
    args = parser.parse_args()
    
    data = []
    for sample in args.samples:
        row = {"Sample": sample}
        
        # Process raw FASTQ
        if args.fastq:
            fastq_file = find_matching_file_or_by_order(sample, args.fastq)
            if args.fastq_report:
                row["Raw Reads"] = extract_read_count_report(fastq_file) if fastq_file else "NA"
            else:
                row["Raw Reads"] = extract_read_count(fastq_file) if fastq_file else "NA"
        
        # Process trimmed FASTQ
        if args.trimmed_fastq:
            trimmed_file = find_matching_file_or_by_order(sample, args.trimmed_fastq)
            if args.fastq_report:
                row["Trimmed Reads"] = extract_read_count_report(trimmed_file) if trimmed_file else "NA"
            else:
                row["Trimmed Reads"] = extract_read_count(trimmed_file) if trimmed_file else "NA"
        
        # Process duplication metrics
        if args.metrics:
            metrics_file = find_matching_file_or_by_order(sample, args.metrics)
            if metrics_file:
                dup_metrics = parse_duplication_metrics(metrics_file)
                row.update(dup_metrics)
            else:
                row.update({"Unpaired Reads Examined": "NA","Read Pairs Examined": "NA", "Secondary or supplementary RDS": "NA", "Unmapped Reads": "NA","Unpaired Read Duplicates": "NA","Read Pair Duplicates": "NA", "Percent Duplication": "NA", })
        
        # Process peaks
        if args.peaks:
            peak_file = find_matching_file(sample, args.peaks)
            row["Number of Peaks"] = count_peaks(peak_file) if peak_file else "NA"
            row["Peak Length Quantiles"] = calculate_peak_length_quantiles(peak_file) if peak_file else "NA"
        
        data.append(row)
    
    # Create DataFrame and save
    df = pd.DataFrame(data)
    df['kept'] = df['Trimmed Reads']/df['Raw Reads']*100
    df['Unmapped Reads'] = (1 - df['Unmapped Reads']/df['Trimmed Reads']/2)*100
    df['Unmapped Reads'] = df['Unmapped Reads'].apply(lambda x: f'{x:.4g}') + '%'
    df = df.rename(columns={'Unmapped Reads':'Mapping Rate','Trimmed Reads':'Kept Trimmed Reads','Unpaired Reads Examined':'Unpaired Reads Aligned','Read Pairs Examined':'Read Pairs Aligned'})
    df['single_rate'] = df['Unpaired Reads Aligned']/df['Kept Trimmed Reads']/2*100
    df['pair_rate'] = df['Read Pairs Aligned']/df['Kept Trimmed Reads']*100
    df['single_dup_rate'] = df['Unpaired Read Duplicates'] / df['Unpaired Reads Aligned']*100
    df['pair_dup_rate'] = df['Read Pair Duplicates']/df['Read Pairs Aligned']*100
    df['Kept Trimmed Reads'] = df['Kept Trimmed Reads'].astype(str) + '(' + df['kept'].apply(lambda x: f'{x:.4g}') + '%)' 
    df['Unpaired Reads Aligned'] = df['Unpaired Reads Aligned'].astype(str) + '(' + df['single_rate'].apply(lambda x: f'{x:.4g}') + '%)'
    df['Read Pairs Aligned'] = df['Read Pairs Aligned'].astype(str) + '(' + df['pair_rate'].apply(lambda x: f'{x:.4g}') + '%)'
    df['Unpaired Read Duplicates'] = df['Unpaired Read Duplicates'].astype(str) + '(' + df['single_dup_rate'].apply(lambda x: f'{x:.4g}') + '%)'  
    df['Read Pair Duplicates'] = df['Read Pair Duplicates'].astype(str) + '(' + df['pair_dup_rate'].apply(lambda x: f'{x:.4g}') + '%)'  
    df = df.drop(columns=['kept','single_rate','pair_rate','single_dup_rate','pair_dup_rate'])
    if args.peaks:
        df = df[['Sample', 'Raw Reads', 'Kept Trimmed Reads',  'Mapping Rate', 'Unpaired Reads Aligned', 'Read Pairs Aligned', 'Secondary or supplementary RDS', 'Unpaired Read Duplicates', 'Read Pair Duplicates', 'Percent Duplication', 'Number of Peaks','Peak Length Quantiles']]
    else:
        df = df[['Sample', 'Raw Reads', 'Kept Trimmed Reads',  'Mapping Rate', 'Unpaired Reads Aligned', 'Read Pairs Aligned', 'Secondary or supplementary RDS', 'Unpaired Read Duplicates', 'Read Pair Duplicates', 'Percent Duplication']]
    df.to_csv(args.output, sep='\t', index=False)
    excel_path=args.output+'.xlsx'
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        # Access the workbook and sheet
        worksheet = writer.sheets['Sheet1']
        # Auto-adjust column widths
        for i, col in enumerate(df.columns):
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[get_column_letter(i + 1)].width = max_len + 2
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    print(f"Summary saved to {args.output}")

if __name__ == "__main__":
    main()
